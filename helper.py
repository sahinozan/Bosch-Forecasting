from typing import Any
import os
import glob
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
import openpyxl
from openpyxl import styles
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import timedelta
from pmdarima import auto_arima
import statsmodels.api as sm
from sklearn.model_selection import train_test_split
from statsmodels.api import tsa
from math import sqrt
import tensorflow as tf
from keras.layers import Dense, LSTM, Dropout
from keras.models import Sequential
from sklearn.metrics import mean_squared_error


def select_pipe_and_split_data(data: pd.DataFrame,
                               pipe_index: int = 0,
                               test_size: float = 0.2,
                               split_index: int = -9,
                               is_sklearn: bool = False) -> tuple[Any, pd.Series, pd.Series]:
    """
    Selects the pipe number and splits the data into train and test.

    Args:
        data: The dataframe to be split.
        pipe_index: The index of the pipe number.
        test_size: The size of the test data.
        split_index: The index to split the data.
        is_sklearn: Whether to use sklearn to split the data.

    Returns:
        The pipe number, the train data, and the test data.
    """
    # get the pipe number
    pipe_number = data.columns[pipe_index]

    # get the data for the given pipe number
    pipe_data = data[pipe_number]

    if is_sklearn:
        # split the data into train and test using sklearn
        train, test = train_test_split(pipe_data, test_size=test_size, shuffle=False, random_state=42)
    else:
        # split the data into train and test without sklearn
        train, test = pipe_data[:split_index], pipe_data[split_index:]

    return pipe_number, train, test


def create_arima_model(train_data: pd.Series,
                       test_data: pd.Series,
                       out_of_sample_size: int = 15,
                       verbose: bool = False,
                       operation_type: str = 'prediction',
                       model_type: str = 'arima') -> tuple[pd.Series, float]:
    """
    Creates and runs an ARIMA model on the given data.

    Args:
        train_data: The data to train the model on.
        test_data: The data to test the model on.
        out_of_sample_size: The number of samples to predict.
        verbose: Whether to print the model summary. 
        operation_type: The type of operation to run.
        model_type: The type of model to run.
    Returns:
        The predicted values and the RMSE.
    """

    # find the best parameters for the model (total: 77, test: 68, train: 9)
    auto_arima_model = auto_arima(train_data,
                                  m=12,
                                  start_P=0, seasonal=True,
                                  d=1, D=1, trace=False,
                                  error_action='ignore',
                                  suppress_warnings=True,
                                  random_state=42, n_fits=50,
                                  maxiter=1000, n_jobs=-1,
                                  out_of_sample_size=out_of_sample_size,
                                  scoring='mse')

    if model_type == "sarimax":
        model = sm.tsa.statespace.SARIMAX(train_data, order=auto_arima_model.order,
                                          seasonal_order=auto_arima_model.seasonal_order)
    elif model_type == "arima":
        model = tsa.ARIMA(train_data, order=auto_arima_model.order,
                          seasonal_order=auto_arima_model.seasonal_order)
    else:
        raise ValueError("model_type must be either 'arima' or 'sarimax'")

    # fit the model
    if model_type == "arima":
        model_fit = model.fit()
    else:
        if verbose:
            model_fit = model.fit(disp=1)
        else:
            model_fit = model.fit(disp=0)

    if operation_type == "forecast":
        # fit the model
        predictions = model_fit.get_forecast(steps=12,
                                             alpha=0.05,
                                             freq='W-MON')
        # calculate root mean squared error
        rmse = sqrt(mean_squared_error(test_data, predictions.predicted_mean))
    elif operation_type == "prediction":
        # fit the model
        predictions = model_fit.predict(start=len(train_data),
                                        end=len(train_data) + len(test_data) - 1,
                                        dynamic=False,
                                        typ='levels').rename('SARIMA Predictions')

        # calculate root mean squared error
        rmse = sqrt(mean_squared_error(test_data, predictions))
    else:
        raise ValueError("operation_type must be either 'forecast' or 'prediction'")

    return predictions, rmse


def create_transposed_and_unique_df(master_df: pd.DataFrame,
                                    sheet_names: list[str],
                                    file_dir) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Transpose the dataframe and create a multi-level index dataframe
    Drop unnecessary columns and rows and write the dataframe to an Excel file

    Args:
        master_df: The dataframe to be transposed
        sheet_names: The names of the sheets
        file_dir: The directory of the Excel file

    Returns:
        A tuple of the dataframe and the multi-level index dataframe (transposed)
    """
    # transpose the dataframe
    master_df_T = master_df.copy().T
    master_df_T.columns, master_df_T.loc["X", :] = master_df_T.loc[["X"], :].values[0], master_df_T.columns

    # create multilevel columns
    exp_df = pd.DataFrame(columns=pd.MultiIndex.from_product([master_df_T.columns, ["Pipe TTNr", "Total"]]).unique())

    # add the data to the multilevel columns
    for i in range(len(master_df_T)):
        exp_df.loc[i, :] = master_df_T.iloc[i, :].values

    # drop the unnecessary index
    exp_df = exp_df.copy().drop(index=0, inplace=False)

    write_to_excel(file_dir=file_dir, df=exp_df, sheet_names=sheet_names)

    # create multi-level index
    three_level_columns = create_three_level_index(df=exp_df)

    # # set the multi-level index
    exp_df_th = exp_df.copy()
    exp_df_th.columns = pd.MultiIndex.from_tuples(three_level_columns)

    return exp_df, exp_df_th


def write_to_excel(file_dir: str,
                   df: pd.DataFrame,
                   sheet_names=None) -> None:
    """
    Writes two versions of the dataframe to an Excel file (one transposed and one not)

    Args:
        file_dir: The directory of the Excel file
        df: The dataframe to be written
        sheet_names: The names of the sheets
    """
    if sheet_names is None:
        sheet_names = ["General", "Experimental"]
    with pd.ExcelWriter(file_dir, engine="openpyxl", mode="w") as writer:
        df.to_excel(writer, sheet_name=sheet_names[0])
        df.T.to_excel(writer, sheet_name=sheet_names[1])


def create_unique_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Transpose the dataframe and get the unique values of the Pipe TTNr column

    Args:
        df: The dataframe to be transposed

    Returns:
        A dataframe with the unique values of the Pipe TTNr column
    """

    vertical_df = pd.DataFrame()

    for i in range(0, len(df.columns), 2):
        temp = df.copy().iloc[:, i:i + 2]
        temp.columns = ['Pipe TTNr', 'Total']
        vertical_df = pd.concat([vertical_df, temp], axis=0, ignore_index=True)

    # get the rows with unique pipes
    unique_pipe_rows = vertical_df["Pipe TTNr"].unique()

    # add the total column for the rows with same Pipe TTNr
    for i in range(len(unique_pipe_rows)):
        vertical_df.loc[vertical_df["Pipe TTNr"] == unique_pipe_rows[i], "Total"] = \
            vertical_df.loc[vertical_df["Pipe TTNr"] == unique_pipe_rows[i], "Total"].sum()

    # drop duplicate Pipe TTNrs
    final_df = vertical_df.drop_duplicates(subset="Pipe TTNr",
                                           keep="first",
                                           inplace=False,
                                           ignore_index=True).sort_values(by="Total", ascending=False).copy()
    final_df = final_df.reset_index(drop=True, inplace=False).copy()

    final_df['Pipe TTNr'] = final_df['Pipe TTNr'].astype(str)

    return final_df


def get_file_indexes(file_dict: dict) -> dict:
    """
    Get the file indexes from the dictionary

    Args:
        file_dict: A dictionary with the year as the key and the files as the value

    Returns:
        A list of the file indexes
    """
    available_file_indexes = [{a: [c.split("_")[0][2:] for c in sorted(b)]} for a, b in file_dict.items()]
    available_file_indexes = {k: v for d in available_file_indexes for k, v in d.items()}

    return available_file_indexes


def create_file_dict(file_dict: dict,
                     years=None) -> dict:
    """
    Get the files from the given years (filtering)

    Args:
        file_dict: A dictionary with the year as the key and the files as the value
        years: The years to be filtered

    Returns:
        A dictionary with the year as the key and the files as the value
    Args:
    """
    # sort the file_dict by the file name number
    if years is None:
        years = [2022, 2023]
    for year, files in file_dict.items():
        file_dict[year] = sorted(files, key=lambda x: int(x.split("_")[0][2:]))

    file_in_given_years = {}

    # get the files from the given years
    if len(years) > 1:
        file_in_given_years = {year: file_dict[year] for year in years}
    elif len(years) == 1:
        file_in_given_years = {years[0]: file_dict[years[0]]}
    elif len(years) == 0:
        raise ValueError("The years list is empty")

    return file_in_given_years


def combine_all_files_within_threshold(file_dict: dict,
                                       master_dir: str,
                                       threshold_df: pd.DataFrame) -> pd.DataFrame:
    """
    Combine all files within the given threshold into a single dataframe

    Args:
        file_dict: A dictionary with the year as the key and the files as the value
        master_dir: The master directory of the files
        threshold_df: A dataframe with the pipes that are within the threshold

    Returns:
        A dataframe with all the files within the threshold
    """
    all_in_one = pd.DataFrame()

    for year, plan in file_dict.items():
        for file in plan:
            df = pd.read_excel(f'{master_dir}/{str(year)}/{file}', sheet_name='Pivot')
            df = df.fillna(0)
            df.iloc[:, 3:26] = df.iloc[:, 3:26].apply(pd.to_numeric, errors='coerce')

            df['Total'] = df.iloc[:, 3:26].sum(axis=1)
            df.loc["Hat", "Total"] = 0
            df = df.sort_values(by=['Total'], inplace=False, ascending=False, ignore_index=True).copy()

            temp = df.iloc[:, [1, -1]].copy()
            index_name = temp.columns[0].split(" ")[3]
            temp.index.name = index_name

            temp.columns = ['Pipe TTNr', temp.columns[-1]]

            temp['Pipe TTNr'] = temp['Pipe TTNr'].astype(str)

            # drop the rows with NaN or non-numeric values
            temp = temp.loc[temp["Pipe TTNr"].apply(lambda x: x.isnumeric()), :].copy()

            temp['Pipe TTNr'] = temp['Pipe TTNr'].astype(int)

            # filter the pipes that are in the 70
            temp = temp.loc[temp["Pipe TTNr"].isin(threshold_df)].copy()

            temp = temp.groupby('Pipe TTNr').sum().copy()

            temp.reset_index(inplace=True)

            # add another column level to aa
            try:
                temp.columns = pd.MultiIndex.from_product(
                    [[pd.to_datetime(index_name, format='%d.%m.%Y').date()], temp.columns])
            except ValueError:
                temp.columns = pd.MultiIndex.from_product(
                    [[pd.to_datetime(index_name, format='%Y-%m-%d').date() + timedelta(days=1)], temp.columns])

            all_in_one = pd.concat([all_in_one, temp], axis=1, ignore_index=False)

    return all_in_one


def convert_multi_to_single_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Converts a multi-index dataframe to a single index dataframe

    Args:
        df: The dataframe to be converted

    Returns:
        A single index dataframe
    """
    df = df.reset_index(inplace=False).copy()
    df = df.rename(columns={"index": "Date"}, inplace=False).copy()

    # if there are NaN values, drop those columns
    df = df.dropna(axis=1, inplace=False)

    # drop all_in_one.columns[1::2] and make it a single column
    pipes = df.iloc[:, 1]
    df = df.drop(df.columns[1::2], axis=1, inplace=False)

    # # insert the column as the first column
    df.insert(1, 'Pipe TTNr', pipes)

    # drop the first column
    df = df.drop(df.columns[0], axis=1, inplace=False)

    # drop the second level of the columns
    df.columns = df.columns.droplevel(1)

    # transpose the dataframe
    df_T = df.T

    # make the first row the column names
    df_T.columns = df_T.iloc[0]
    df_T = df_T.drop(df_T.index[0])

    # convert all the column names to str
    df_T.columns = df_T.columns.astype(int).astype(str)

    return df_T


def get_occurrences_per_file(df: pd.DataFrame) -> pd.DataFrame:
    """
    Gets the occurrences of each pipe in each file

    Args:
        df: The dataframe to be checked for occurrences

    Returns:
        A dataframe with the occurrences of each pipe in each file
    """

    pipes_dict = {}

    for i in range(0, len(df.columns), 2):
        temp = df.copy().iloc[:, i:i + 2]
        pipes_dict[temp.columns[0][0]] = temp[(temp.columns[0][0], 'Pipe TTNr')].unique()

    # count the number of occurrences for each pipe and which weeks it occurs
    pipe_occurrences, week_occurrences = {}, {}

    for key, value in pipes_dict.items():
        for pipe in value:
            if pipe in pipe_occurrences.keys():
                pipe_occurrences[pipe] += 1
                week_occurrences[pipe].append(key)
            else:
                pipe_occurrences[pipe] = 1
                week_occurrences[pipe] = [key]

    merged_dict = {k: [pipe_occurrences[k], week_occurrences[k]] for k in pipe_occurrences.keys()}

    # create a dataframe from the dictionary
    pipe_occurrences_df = pd.DataFrame.from_dict(merged_dict, orient='index', columns=['Occurrences', 'Weeks'])
    pipe_occurrences_df = pipe_occurrences_df.sort_values(by="Occurrences", ascending=False, inplace=False)

    return pipe_occurrences_df


def get_occurrences_with_threshold(df: pd.DataFrame,
                                   threshold: int = 50) -> list[int]:
    """
    Gets the occurrences of each pipe in each file with a threshold

    Args:
        df: The dataframe to be checked for an occurrence
        threshold: The threshold for the occurrences (default: 50)

    Returns:
        A dataframe with the occurrences of each pipe in each file with a threshold
    """
    pipes_in_threshold_df = df.loc[df["Occurrences"] >= threshold, :].sort_values(by="Occurrences",
                                                                                  ascending=False)

    pipes_in_threshold_df = pipes_in_threshold_df.reset_index(inplace=False).copy()
    pipes_in_threshold_df = pipes_in_threshold_df.rename(columns={"index": "Pipe TTNr"}, inplace=False).copy()

    pipes_in_threshold_df = pipes_in_threshold_df['Pipe TTNr'].values
    pipes_in_threshold_df = [int(x) for x in pipes_in_threshold_df]

    return pipes_in_threshold_df


def find_common_pipes(file_index: int,
                      file_year: int,
                      top_level_df: pd.DataFrame,
                      file_dict: dict[int, list[str]],
                      master_dir: str,
                      threshold: int = 20
                      ) -> pd.DataFrame:
    """
    Finds the most produced pipes for the selected production plan

    Args:
        file_index: The file index (4, 9, 27, 36, 41,...)
        file_year: The file year (2021, 2022, 2023)
        top_level_df: The top level dataframe
        file_dict: The file dictionary
        master_dir: The master directory
        threshold: The number of pipes to be selected

    Returns:
        The top level dataframe with the most produced pipes for the selected production plan
    """
    df = pd.read_excel(f'{master_dir}/{str(file_year)}/{file_dict[file_year][file_index]}', sheet_name='Pivot')

    df.iloc[:, 3:26] = df.iloc[:, 3:26].apply(pd.to_numeric, errors='coerce')

    # combine the rows with the same pipe code
    df = df.groupby(df.iloc[:, 1]).sum()

    df['Total'] = df.iloc[:, 3:26].sum(axis=1)
    df.loc["Hat", "Total"] = 0

    # sort the dataframe by the Total column
    df.sort_values(by=['Total'], inplace=True, ascending=False)

    # transpose the dataframe and select the top 20 pipes
    top_20_pipes = df.iloc[:threshold, [1, -1]].T

    # replace the first row with the column values
    top_20_pipes.iloc[0, :], top_20_pipes.columns = pd.Series(
        map(str, top_20_pipes.columns)), list(range(1, threshold + 1))

    # rename the columns
    top_20_pipes.index = pd.Index(['Pipe TTNr', 'Total'])

    # top_20_pipes.iloc[0, :] = top_20_pipes.iloc[0, :].astype(int)

    # insert an empty column to first position
    top_20_pipes.insert(0, "X", file_dict[file_year][file_index].split("_")[0] + f"_{file_year}")

    # add the top 20 pipes to the top_level_df
    top_level_df = pd.concat([top_level_df, top_20_pipes], axis=0)

    return top_level_df


def configure_matplotlib(labelsize: int = 18,
                         titlesize: int = 22,
                         titlepad: int = 25,
                         labelpad: int = 15,
                         tick_major_pad: int = 10,
                         dpi: int = 200) -> None:
    """
    Configures matplotlib to use the fivethirtyeight style and the Ubuntu font.
    Args:
        labelsize: The size of the axis labels
        titlesize: The size of the title
        titlepad: The padding of the title
        labelpad: The padding of the axis labels
        tick_major_pad: The padding of the major ticks
        dpi: The resolution of the figure
    """
    plt.rcParams['font.family'] = 'Arial'
    plt.style.use('fivethirtyeight')
    plt.rcParams['font.serif'] = 'Ubuntu'
    plt.rcParams['font.monospace'] = 'Ubuntu Mono'
    plt.rcParams['axes.labelweight'] = 'bold'
    plt.rcParams['axes.labelsize'] = labelsize
    plt.rcParams['axes.labelpad'] = labelpad
    plt.rcParams['axes.titlesize'] = titlesize
    plt.rcParams['axes.titleweight'] = 'bold'
    plt.rcParams['axes.titlepad'] = titlepad
    plt.rcParams['figure.dpi'] = dpi
    plt.rcParams['xtick.major.pad'] = tick_major_pad
    plt.rcParams['ytick.major.pad'] = tick_major_pad

    # change the background color of the figure
    plt.rcParams['figure.facecolor'] = 'none'
    plt.rcParams['axes.facecolor'] = 'none'

    # change the color of the grid
    plt.rcParams['axes.grid'] = False


def rgb_to_hex(r, g, b) -> str:
    """
    Converts an RGB color to hex. (255,255,255 -> FFFFFF)
    Args:
        r: The red value
        g: The green value
        b: The blue value

    Returns:
        The hex string
    """
    return '#{:02x}{:02x}{:02x}'.format(r, g, b)


def hex_to_RGB(hex_str) -> list[int]:
    """
    Converts a hex string to RGB. (FFFFFF -> [255,255,255])
    Args:
        hex_str: The hex string

    Returns:
        A list of RGB values
    """
    return [int(hex_str[x:x + 2], 16) for x in range(1, 6, 2)]


def get_color_gradient(c1, c2, n) -> list[str]:
    """
    Given two hex colors, returns a color gradient
    with n colors.

    Args:
        c1: The first color (hex)
        c2: The second color (hex)
        n: The number of colors in the gradient

    Returns:
        A list of hex colors
    """
    assert n > 1
    c1_rgb = np.array(hex_to_RGB(c1)) / 255
    c2_rgb = np.array(hex_to_RGB(c2)) / 255
    mix_pcts = [x / (n - 1) for x in range(n)]
    rgb_colors = [((1 - mix) * c1_rgb + (mix * c2_rgb)) for mix in mix_pcts]
    return ["#" + "".join([format(int(round(val * 255)), "02x") for val in item]) for item in rgb_colors]


def create_bar_plot(df: pd.DataFrame,
                    selected_year: int,
                    file_index: str,
                    ascending: bool = True,
                    threshold: int = 20) -> None:
    """
    Creates a bar plot for the selected year and file index

    Args:
        df: The dataframe that contains the data (exp_df)
        selected_year: The selected year (2021, 2022, 2023)
        file_index: The file index (4, 9, 27, 36, 41,...)
        ascending: If the plot should be sorted ascending or descending
        threshold: The number of pipes to be selected
    """
    if len(str(file_index)) == 1:
        file_index = "0" + str(file_index)

    selected_file = f'KW{str(file_index)}_{selected_year}'

    if (str(selected_year), selected_file, "Pipe TTNr") not in df.columns:
        print(f">>> KW{str(file_index)} does not exist for the year {selected_year}!")
        return None

    configure_matplotlib()
    fig, ax = plt.subplots(figsize=(20, 10))

    # sorted plot
    sns.barplot(ax=ax,
                data=df,
                x=df[(str(selected_year), selected_file, "Pipe TTNr")],
                y=df[(str(selected_year), selected_file, "Total")],
                order=df.sort_values(by=(str(selected_year), selected_file, "Total"), ascending=ascending).head(
                    threshold)[(str(selected_year), selected_file, "Pipe TTNr")],

                color="blue",
                hue_order=df.sort_values(by=(str(selected_year), selected_file, "Total"), ascending=ascending).head(
                    threshold)[(str(selected_year), selected_file, "Pipe TTNr")],
                label=selected_file,
                errorbar=None,
                palette=get_color_gradient("#1f77b4", "#ff7f0e", threshold))

    sns.despine(fig=fig, ax=ax, top=True, right=True, left=True, bottom=True)

    # rotate the x-ticks
    plt.xticks(rotation=90)

    # add padding to x-ticks
    plt.tick_params(axis='x', which='major', pad=10)

    plt.xlabel("Pipe TTNr", labelpad=25)
    plt.ylabel("Total Quantity", labelpad=25)
    plt.title(f"Top {threshold} Pipes in {selected_file}", pad=25)

    plt.show()


def unique_pipe_bar_plot(pipe_df: pd.DataFrame,
                         total_quantity_limit: int,
                         fig_size: tuple = (16, 20),
                         rotation: str = 'horizontal',
                         ascending: bool = True,
                         years=None,
                         threshold: int = 20) -> None:
    """
    Creates a bar plot for the unique pipes
    Args:
        pipe_df: The dataframe that contains the data (final_df)
        total_quantity_limit: The minimum total quantity limit
        fig_size: The figure size
        rotation: The rotation of the x-axis labels
        ascending: If the plot should be sorted ascending or descending
        years: The years to be considered
        threshold: The number of pipes to be selected
    """
    if years is None:
        years = [2022, 2023]
    configure_matplotlib()
    fig, ax = plt.subplots(figsize=fig_size)

    if rotation == 'horizontal':
        # horizontal plot (better with more labels)
        ax = sns.barplot(y="Pipe TTNr",
                         x="Total",
                         data=pipe_df.loc[np.logical_and(pipe_df["Pipe TTNr"].apply(lambda x: x.isnumeric()),
                                                         pipe_df["Total"] > total_quantity_limit), :].copy(),
                         order=pipe_df.loc[np.logical_and(pipe_df["Pipe TTNr"].apply(lambda x: x.isnumeric()),
                                                          pipe_df["Total"] > total_quantity_limit), :].copy(

                         ).sort_values(
                             by="Total", ascending=ascending).head(threshold)["Pipe TTNr"],
                         color="blue",
                         errorbar=None,
                         palette=get_color_gradient("#1f77b4", "#ff7f0e", threshold),
                         hue_order=pipe_df.loc[np.logical_and(pipe_df["Pipe TTNr"].apply(lambda x: x.isnumeric()),
                                                              pipe_df["Total"] > total_quantity_limit), :].copy()[
                             "Pipe TTNr"].unique()
                         )

        plt.ylabel("Pipe TTNr", labelpad=25)
        plt.xlabel("Total Quantity", labelpad=25)

        # add grid lines
        ax.grid(axis="x", color="black", linestyle="dashed", linewidth=0.5)

        # add padding to y-ticks
        plt.tick_params(axis='x', which='major', pad=10)
    elif rotation == 'vertical':
        # vertical plot (better with fewer labels)
        ax = sns.barplot(x="Pipe TTNr",
                         y="Total",
                         data=pipe_df.loc[np.logical_and(pipe_df["Pipe TTNr"].apply(lambda x: x.isnumeric()),
                                                         pipe_df["Total"] > total_quantity_limit), :].copy(),
                         order=pipe_df.loc[np.logical_and(pipe_df["Pipe TTNr"].apply(lambda x: x.isnumeric()),
                                                          pipe_df["Total"] > total_quantity_limit), :].copy(

                         ).sort_values(
                             by="Total", ascending=ascending).head(threshold)["Pipe TTNr"],
                         color="blue",
                         errorbar=None,
                         palette=get_color_gradient("#1f77b4", "#ff7f0e", threshold),
                         hue_order=pipe_df.loc[np.logical_and(pipe_df["Pipe TTNr"].apply(lambda x: x.isnumeric()),
                                                              pipe_df["Total"] > total_quantity_limit), :].copy()[
                             "Pipe TTNr"].unique()
                         )
        plt.xlabel("Pipe TTNr", labelpad=25)
        plt.ylabel("Total Quantity", labelpad=25)

        # add grid lines
        ax.grid(axis="y", color="black", linestyle="dashed", linewidth=0.5)

        # add padding to y-ticks
        plt.tick_params(axis='y', which='major', pad=10)
    else:
        raise ValueError("rotation must be either 'horizontal' or 'vertical'")

    sns.despine(fig=fig, ax=ax, top=True, right=True, left=True, bottom=True)

    # rotate the x-ticks
    plt.xticks(rotation=90, fontsize=12)

    # set the font size of the y-ticks
    plt.yticks(fontsize=12)

    plt.title(f"Overall Top Pipes {years[0]}-{years[-1]}", pad=25)

    plt.show()


def format_general_sheet(file_dir: str) -> None:
    """
    Format the Excel file. This formatting involves removing the blank rows, setting the column header,
    setting the column, row width and height and setting the font size and alignment.

    Args:
        file_dir: The file directory of the Excel file
    """
    wb = openpyxl.load_workbook(file_dir)
    if "General" in wb.sheetnames:
        ws = wb["General"]
    else:
        print(">>> The sheet 'General' does not exist!")
        return None

    # remove the blank rows
    ws.delete_rows(3)

    # set the column header
    ws['A2'] = "Ranking"

    # set the column width
    for i in range(1, ws.max_column + 1):
        if i % 2 == 0:
            ws.column_dimensions[get_column_letter(i)].width = 15  # type: ignore
        elif i % 2 == 1:
            ws.column_dimensions[get_column_letter(i)].width = 10  # type: ignore

    # set the row height
    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = 20  # type: ignore

    # set the font size and alignment
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            ws.cell(row=i, column=j).font = Font(size=10)
            ws.cell(row=i, column=j).alignment = Alignment(horizontal='center', vertical='center')

    # set the column header font size and font bold
    for i in range(1, ws.max_column + 1):
        ws.cell(row=1, column=i).font = Font(size=12, bold=True)
        ws.cell(row=2, column=i).font = Font(size=10, bold=True)

    for i in range(1, ws.max_row + 1):
        ws.cell(row=i, column=1).font = Font(size=10, bold=True)

    # create the borders
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            ws.cell(row=i, column=j).border = Border(
                left=styles.borders.Side(border_style='thin', color='000000'),
                right=styles.borders.Side(border_style='thin',
                                          color='000000'),
                top=styles.borders.Side(border_style='thin', color='000000'),
                bottom=styles.borders.Side(border_style='thin',
                                           color='000000'))

    wb.save(file_dir)


def format_experimental_sheet(file_dir: str) -> None:
    """
    Format the Excel file. This formatting involves removing the blank rows, setting the column header,
    setting the column, row width and height and setting the font size and alignment.

    Args:
        file_dir: The file directory of the Excel file
    """
    wb = openpyxl.load_workbook(file_dir)
    if "Experimental" in wb.sheetnames:
        ws = wb["Experimental"]
    else:
        print(">>> The sheet 'Experimental' does not exist!")
        return None

    # set the column header
    ws['A1'] = "File Index"
    ws['B1'] = "Ranking"

    # set the column width
    for i in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15  # type: ignore

    # set the column of the file index column
    ws.column_dimensions[get_column_letter(1)].width = 20  # type: ignore

    # set the row height
    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = 20  # type: ignore

    # set the font size and alignment
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            ws.cell(row=i, column=j).font = Font(size=10)
            ws.cell(row=i, column=j).alignment = Alignment(horizontal='center', vertical='center')

    # set the column header font size and font bold
    for i in range(1, ws.max_column + 1):
        ws.cell(row=1, column=i).font = Font(size=10, bold=True)

    for i in range(1, ws.max_row + 1):
        ws.cell(row=i, column=1).font = Font(size=12, bold=True)
        ws.cell(row=i, column=2).font = Font(size=10, bold=True)

    # create the borders
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            ws.cell(row=i, column=j).border = Border(
                left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin',
                           color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin',
                            color='000000'))

    wb.save(file_dir)


def create_three_level_index(df: pd.DataFrame) -> list[tuple[str | Any, ...]]:
    """
    Creates a three-level index for the dataframe

    Args:
        df: The dataframe that contains the data (exp_df)

    Returns:
        The three-level index
    """
    final_columns = []

    for v, i in enumerate(df.columns):
        if i[0].split("_")[-1] == "2021":
            w = list(i)
            w.insert(0, "2021")
        elif i[0].split("_")[-1] == "2022":
            w = list(i)
            w.insert(0, "2022")
        else:
            w = list(i)
            w.insert(0, "2023")
        final_columns.append(tuple(w))

    return final_columns


def get_data_files() -> tuple[dict[int, list[str]], str, str]:
    """
    Gets the data files and the file directory

    Returns:
        The data files, the file directory and the master directory
    """
    file_dict = {}

    file_dir = "/Users/ozansahin/Downloads/master.xlsx"
    master_dir = "/Users/ozansahin/Downloads/New_Converted_Data"

    for i in [2021, 2022, 2023]:
        os.chdir(f"{master_dir}/{i}")
        extension = 'xlsx'
        file_dict[i] = [i for i in glob.glob('*.{}'.format(extension))]

    return file_dict, file_dir, master_dir
